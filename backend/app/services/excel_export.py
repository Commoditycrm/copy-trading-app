"""Excel (.xlsx) export helpers.

One workbook builder shared by every export endpoint (user trades, admin
trades, admin performance) so all three look and behave the same — frozen
header, auto-filter, sensible widths, real typed cells.

Typed cells matter here. If you hand Excel a string it will guess, and it
guesses badly on exactly the data we export: option symbols like "SPXW
251219C00500000" become scientific notation, and ISO timestamps stay text
so they can't be sorted or filtered as dates. So we pass through real
``datetime`` / ``Decimal`` values and let a number format render them.

Uses ``write_only`` mode: rows stream to the file rather than living in an
in-memory sheet, which keeps a large export from ballooning the worker.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Any, Callable, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

# Excel has no timezone concept — a tz-aware datetime raises on write. We
# normalise everything to UTC and say so in the header, rather than writing
# local times that silently mean different things per reader.
_DATETIME_FMT = "yyyy-mm-dd hh:mm:ss"
_DATE_FMT = "yyyy-mm-dd"
_MONEY_FMT = "#,##0.00######"      # keeps option premiums / fractional shares
_INT_FMT = "#,##0"

_HEADER_FILL = PatternFill("solid", fgColor="1F3350")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)


@dataclass(frozen=True)
class Column:
    """One spreadsheet column.

    ``get`` pulls the raw value off a row object. Return a native type —
    datetime/Decimal/int/str/None — NOT a preformatted string, or Excel
    loses the ability to sort and filter it.
    """
    header: str
    get: Callable[[Any], Any]
    width: int = 16
    fmt: str | None = None


def _clean(value: Any) -> Any:
    """Coerce a value into something openpyxl can write."""
    if value is None:
        return None
    if isinstance(value, datetime):
        # openpyxl raises on tz-aware datetimes. Convert to UTC, then drop
        # the tzinfo — every timestamp we export is UTC by convention.
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)
        return value
    if isinstance(value, (date, Decimal, int, float, bool)):
        return value
    return str(value)


def build_workbook(
    *,
    columns: Sequence[Column],
    rows: Iterable[Any],
    sheet_title: str = "Export",
    meta: Sequence[tuple[str, Any]] = (),
) -> bytes:
    """Render ``rows`` into an .xlsx and return the bytes.

    ``meta`` renders as a short key/value preamble above the table — we use it
    to record which filters produced the file. Without that, an exported
    sheet is indistinguishable from a full dump, and someone will eventually
    read a filtered export as if it were everything.
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_title[:31])  # Excel caps sheet names at 31

    header_at = len(meta) + 2 if meta else 1

    # write_only STREAMS the sheet, so anything structural has to be set before
    # the first append — set it afterwards and openpyxl drops it silently, no
    # error, and you only notice the header doesn't freeze when you open the
    # file. Applies to freeze_panes and column widths alike. (auto_filter is
    # fine after, and has to be: it needs the final row count.)
    ws.freeze_panes = f"A{header_at + 1}"
    dim = DimensionHolder(worksheet=ws)
    for i, col in enumerate(columns, start=1):
        letter = get_column_letter(i)
        dim[letter] = ColumnDimension(ws, index=letter, width=col.width)
    ws.column_dimensions = dim

    if meta:
        for key, value in meta:
            cell_key = _new_cell(ws, key, bold=True)
            ws.append([cell_key, _new_cell(ws, value)])
        ws.append([])

    header_row = []
    for col in columns:
        cell = _new_cell(ws, col.header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        header_row.append(cell)
    ws.append(header_row)

    n = 0
    for row in rows:
        out = []
        for col in columns:
            cell = _new_cell(ws, _clean(col.get(row)))
            if col.fmt:
                cell.number_format = col.fmt
            out.append(cell)
        ws.append(out)
        n += 1

    if n:
        last = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A{header_at}:{last}{header_at + n}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _new_cell(ws, value, *, bold: bool = False):
    from openpyxl.cell import WriteOnlyCell  # noqa: PLC0415
    cell = WriteOnlyCell(ws, value=_clean(value))
    if bold:
        cell.font = Font(bold=True)
    return cell


def filename(prefix: str, *, when: datetime) -> str:
    """Stable, sortable download name: kopyya-trades-20260715-1432.xlsx.

    Caller passes the timestamp — this module stays clock-free so exports are
    reproducible in tests.
    """
    return f"kopyya-{prefix}-{when.strftime('%Y%m%d-%H%M')}.xlsx"


XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
